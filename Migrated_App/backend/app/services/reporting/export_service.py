"""
Export Service

Handles exporting reports to various formats including PDF, Excel, and CSV.
Provides templating and formatting capabilities for professional report output.
"""

import os
import io
import csv
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
import logging

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

# Excel generation
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.schemas.reports import ReportResponse, ExportFormat

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting reports to various formats"""
    
    def __init__(self, export_dir: str = "/tmp/reports"):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)
    
    async def export_report(self, report: ReportResponse, format: ExportFormat) -> str:
        """
        Export report to specified format
        
        Args:
            report: Report data and metadata
            format: Export format (PDF, Excel, CSV)
            
        Returns:
            File path of exported report
        """
        try:
            if format == ExportFormat.PDF:
                return await self._export_to_pdf(report)
            elif format == ExportFormat.EXCEL:
                return await self._export_to_excel(report)
            elif format == ExportFormat.CSV:
                return await self._export_to_csv(report)
            else:
                raise ValueError(f"Unsupported export format: {format}")
                
        except Exception as e:
            logger.error(f"Error exporting report {report.report_id} to {format}: {str(e)}")
            raise
    
    async def _export_to_pdf(self, report: ReportResponse) -> str:
        """Export report to PDF format"""
        try:
            filename = f"{report.report_id}.pdf"
            filepath = os.path.join(self.export_dir, filename)
            
            # Create PDF document
            doc = SimpleDocTemplate(
                filepath,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            story = []
            
            # Add title
            story.append(Paragraph(report.title, title_style))
            story.append(Spacer(1, 12))
            
            # Add report metadata
            metadata_data = [
                ['Report ID:', report.report_id],
                ['Generated:', report.generated_at.strftime('%d/%m/%Y %H:%M:%S')],
                ['Total Records:', str(report.total_records)]
            ]
            
            if report.parameters:
                for key, value in report.parameters.items():
                    metadata_data.append([f'{key.title()}:', str(value)])
            
            metadata_table = Table(metadata_data, colWidths=[2*inch, 4*inch])
            metadata_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(metadata_table)
            story.append(Spacer(1, 20))
            
            # Add data based on report type
            if report.report_type == 'trial_balance':
                story.extend(self._format_trial_balance_pdf(report.data))
            elif report.report_type == 'profit_loss':
                story.extend(self._format_profit_loss_pdf(report.data))
            elif report.report_type == 'balance_sheet':
                story.extend(self._format_balance_sheet_pdf(report.data))
            elif report.report_type in ['customer_aging', 'supplier_aging']:
                story.extend(self._format_aging_report_pdf(report.data, report.report_type))
            elif report.report_type == 'stock_valuation':
                story.extend(self._format_stock_valuation_pdf(report.data))
            else:
                # Generic table format
                story.extend(self._format_generic_table_pdf(report.data))
            
            # Build PDF
            doc.build(story)
            
            logger.info(f"PDF report exported to: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error creating PDF report: {str(e)}")
            raise
    
    def _format_trial_balance_pdf(self, data: List[Dict]) -> List:
        """Format trial balance data for PDF"""
        story = []
        
        # Create table data
        table_data = [['Account Code', 'Account Name', 'Account Type', 'Debit', 'Credit']]
        
        for row in data:
            if row['account_code'] == 'TOTAL':
                # Add separator line before total
                story.append(Spacer(1, 10))
                
            table_data.append([
                row['account_code'],
                row['account_name'],
                row['account_type'],
                f"£{row['debit_amount']:,.2f}" if row['debit_amount'] > 0 else '',
                f"£{row['credit_amount']:,.2f}" if row['credit_amount'] > 0 else ''
            ])
        
        # Create table
        table = Table(table_data, colWidths=[1.2*inch, 3*inch, 1.5*inch, 1.2*inch, 1.2*inch])
        
        # Apply styles
        style = TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),  # Right align amounts
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.beige, colors.white]),
            
            # Total row styling
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            
            # Borders
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        
        table.setStyle(style)
        story.append(table)
        
        return story
    
    def _format_profit_loss_pdf(self, data: Dict) -> List:
        """Format P&L data for PDF"""
        story = []
        
        # Income section
        story.append(Paragraph('<b>INCOME</b>', getSampleStyleSheet()['Heading2']))
        
        income_data = [['Account Code', 'Account Name', 'Amount']]
        for item in data['income']:
            income_data.append([
                item['account_code'],
                item['account_name'],
                f"£{item['amount']:,.2f}"
            ])
        income_data.append(['', 'Total Income', f"£{data['income_total']:,.2f}"])
        
        income_table = Table(income_data, colWidths=[1.5*inch, 4*inch, 1.5*inch])
        income_table.setStyle(self._get_financial_table_style())
        story.append(income_table)
        story.append(Spacer(1, 20))
        
        # Expenses section
        story.append(Paragraph('<b>EXPENSES</b>', getSampleStyleSheet()['Heading2']))
        
        expense_data = [['Account Code', 'Account Name', 'Amount']]
        for item in data['expenses']:
            expense_data.append([
                item['account_code'],
                item['account_name'],
                f"£{item['amount']:,.2f}"
            ])
        expense_data.append(['', 'Total Expenses', f"£{data['expense_total']:,.2f}"])
        
        expense_table = Table(expense_data, colWidths=[1.5*inch, 4*inch, 1.5*inch])
        expense_table.setStyle(self._get_financial_table_style())
        story.append(expense_table)
        story.append(Spacer(1, 20))
        
        # Net profit
        net_profit_data = [['Net Profit', f"£{data['net_profit']:,.2f}"]]
        net_profit_table = Table(net_profit_data, colWidths=[5.5*inch, 1.5*inch])
        net_profit_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
        ]))
        story.append(net_profit_table)
        
        return story
    
    def _get_financial_table_style(self) -> TableStyle:
        """Get standard table style for financial reports"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
    
    async def _export_to_excel(self, report: ReportResponse) -> str:
        """Export report to Excel format"""
        try:
            filename = f"{report.report_id}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = report.report_type.title()
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add title and metadata
            ws['A1'] = report.title
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            
            row = 3
            ws[f'A{row}'] = 'Report ID:'
            ws[f'B{row}'] = report.report_id
            row += 1
            ws[f'A{row}'] = 'Generated:'
            ws[f'B{row}'] = report.generated_at.strftime('%d/%m/%Y %H:%M:%S')
            row += 1
            
            if report.parameters:
                for key, value in report.parameters.items():
                    ws[f'A{row}'] = f'{key.title()}:'
                    ws[f'B{row}'] = str(value)
                    row += 1
            
            row += 2  # Add space
            
            # Add data
            if isinstance(report.data, list):
                if report.data:
                    # Get headers from first row
                    headers = list(report.data[0].keys())
                    
                    # Write headers
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header.title())
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    row += 1
                    
                    # Write data
                    for data_row in report.data:
                        for col, header in enumerate(headers, 1):
                            value = data_row.get(header, '')
                            
                            # Format numbers
                            if isinstance(value, (int, float)):
                                if 'amount' in header.lower() or 'total' in header.lower() or 'cost' in header.lower():
                                    ws.cell(row=row, column=col, value=value).number_format = '£#,##0.00'
                                else:
                                    ws.cell(row=row, column=col, value=value)
                            else:
                                ws.cell(row=row, column=col, value=value)
                            
                            ws.cell(row=row, column=col).border = border
                        row += 1
                        
            elif isinstance(report.data, dict):
                # Handle dictionary data (like P&L, Balance Sheet)
                self._write_dict_to_excel(ws, report.data, row)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            
            logger.info(f"Excel report exported to: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error creating Excel report: {str(e)}")
            raise
    
    def _write_dict_to_excel(self, ws, data: Dict, start_row: int) -> int:
        """Write dictionary data to Excel worksheet"""
        row = start_row
        
        for section_name, section_data in data.items():
            if isinstance(section_data, list):
                # Write section header
                ws[f'A{row}'] = section_name.title()
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                # Write section data
                if section_data:
                    headers = list(section_data[0].keys())
                    
                    # Write headers
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=header.title()).font = Font(bold=True)
                    row += 1
                    
                    # Write data
                    for item in section_data:
                        for col, header in enumerate(headers, 1):
                            value = item.get(header, '')
                            if isinstance(value, (int, float)):
                                ws.cell(row=row, column=col, value=value).number_format = '£#,##0.00'
                            else:
                                ws.cell(row=row, column=col, value=value)
                        row += 1
                
                row += 1  # Add space between sections
                
            elif isinstance(section_data, (int, float, str)):
                # Simple key-value pair
                ws[f'A{row}'] = section_name.title()
                ws[f'B{row}'] = section_data
                if isinstance(section_data, (int, float)):
                    ws[f'B{row}'].number_format = '£#,##0.00'
                row += 1
        
        return row
    
    async def _export_to_csv(self, report: ReportResponse) -> str:
        """Export report to CSV format"""
        try:
            filename = f"{report.report_id}.csv"
            filepath = os.path.join(self.export_dir, filename)
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                if isinstance(report.data, list) and report.data:
                    fieldnames = list(report.data[0].keys())
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    
                    # Write header
                    writer.writeheader()
                    
                    # Write data
                    for row in report.data:
                        writer.writerow(row)
                        
                elif isinstance(report.data, dict):
                    # Flatten dictionary data for CSV
                    writer = csv.writer(csvfile)
                    
                    # Write metadata
                    writer.writerow(['Report', report.title])
                    writer.writerow(['Generated', report.generated_at.strftime('%d/%m/%Y %H:%M:%S')])
                    writer.writerow([])  # Empty row
                    
                    # Write data sections
                    for section_name, section_data in report.data.items():
                        writer.writerow([section_name.title()])
                        
                        if isinstance(section_data, list) and section_data:
                            headers = list(section_data[0].keys())
                            writer.writerow(headers)
                            
                            for item in section_data:
                                writer.writerow([item.get(header, '') for header in headers])
                                
                        elif isinstance(section_data, (int, float, str)):
                            writer.writerow([section_data])
                        
                        writer.writerow([])  # Empty row between sections
            
            logger.info(f"CSV report exported to: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error creating CSV report: {str(e)}")
            raise